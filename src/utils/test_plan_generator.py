# ==============================================
# Test Plan Document Generator
# ==============================================

from typing import Dict, Any
from datetime import datetime
import os
from pathlib import Path
import tempfile

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from src.models.test_plan_models import TestPlan
from src.utils.logger import get_logger

logger = get_logger(__name__)

class TestPlanDocumentGenerator:
    """
    Generate test plan documents in PDF and Excel formats
    """

    def __init__(self, output_dir: str = None):
        """
        Initialize document generator
        
        Args:
            output_dir: Directory for generated documents (default: system temp)
        """
        if output_dir:
            self.output_dir = Path(output_dir)
            self.output_dir.mkdir(exist_ok=True)
        else:
            self.output_dir = Path(tempfile.gettempdir())
        
        logger.info(f"Test plan document generator initialized at {self.output_dir}")

    def generate_pdf(self, test_plan: TestPlan, jira_ticket_key: str) -> str:
        """
        Generate PDF test plan document

        Args:
            test_plan: TestPlan object
            jira_ticket_key: Jira ticket key

        Returns:
            Path to generated PDF file
        """
        filename = f"TestPlan_{jira_ticket_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = self.output_dir / filename

        logger.info(f"Generating PDF test plan: {filename}")

        try:
            # Create PDF document
            doc = SimpleDocTemplate(
                str(filepath),
                pagesize=letter,
                rightMargin=0.75*inch,
                leftMargin=0.75*inch,
                topMargin=1*inch,
                bottomMargin=0.75*inch
            )

            elements = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1a1a1a'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            elements.append(Paragraph(f"Test Plan: {jira_ticket_key}", title_style))
            elements.append(Spacer(1, 0.2*inch))

            # Overview section
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#2c5aa0'),
                spaceAfter=12,
                spaceBefore=12
            )

            elements.append(Paragraph("Test Plan Overview", heading_style))

            overview_data = [
                ['Jira Ticket', jira_ticket_key],
                ['Generated Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Test Approach', test_plan.test_approach],
                ['Total Scenarios', str(len(test_plan.test_scenarios))],
                ['Confidence Score', f"{test_plan.confidence_score}%"]
            ]

            overview_table = Table(overview_data, colWidths=[2*inch, 4*inch])
            overview_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f0f0f0')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(overview_table)
            elements.append(Spacer(1, 0.3*inch))

            # Test Strategy
            elements.append(Paragraph("Test Strategy", heading_style))
            elements.append(Paragraph(test_plan.strategy, styles['BodyText']))
            elements.append(Spacer(1, 0.3*inch))

            # Test Scenarios Summary
            elements.append(Paragraph("Test Scenarios Summary", heading_style))
            
            scenario_types = {}
            for scenario in test_plan.test_scenarios:
                test_type = scenario.get('test_type', 'Unknown')
                scenario_types[test_type] = scenario_types.get(test_type, 0) + 1
            
            scenario_data = [['Test Type', 'Count']]
            for test_type, count in scenario_types.items():
                scenario_data.append([test_type, str(count)])

            scenario_table = Table(scenario_data, colWidths=[3*inch, 1*inch])
            scenario_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c5aa0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey)
            ]))
            elements.append(scenario_table)

            # Build PDF
            doc.build(elements)
            logger.info(f"PDF test plan generated successfully: {filepath}")

            return str(filepath)

        except Exception as e:
            logger.error(f"Failed to generate PDF test plan: {str(e)}")
            raise

    def generate_excel(self, test_plan: TestPlan, jira_ticket_key: str) -> str:
        """
        Generate Excel test plan workbook

        Args:
            test_plan: TestPlan object
            jira_ticket_key: Jira ticket key

        Returns:
            Path to generated Excel file
        """
        filename = f"TestPlan_{jira_ticket_key}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = self.output_dir / filename

        logger.info(f"Generating Excel test plan: {filename}")

        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Overview sheet
            ws = wb.active
            ws.title = "Overview"
            
            # Add header
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2C5AA0', end_color='2C5AA0', fill_type='solid')
            
            ws['A1'] = f"Test Plan: {jira_ticket_key}"
            ws['A1'].font = Font(name='Calibri', size=14, bold=True)
            
            # Add overview data
            overview_data = [
                ['Jira Ticket', jira_ticket_key],
                ['Generated Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Test Approach', test_plan.test_approach],
                ['Total Scenarios', len(test_plan.test_scenarios)],
                ['Confidence Score', f"{test_plan.confidence_score}%"],
                ['Strategy', test_plan.strategy]
            ]
            
            for row_idx, (key, value) in enumerate(overview_data, 3):
                ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
                ws.cell(row=row_idx, column=2, value=value)
            
            # Test Scenarios sheet
            ws_scenarios = wb.create_sheet("Test Scenarios")
            
            headers = ['ID', 'Title', 'Priority', 'Type', 'Given', 'When', 'Then']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_scenarios.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Add scenarios
            for row_idx, scenario in enumerate(test_plan.test_scenarios, 2):
                ws_scenarios.cell(row=row_idx, column=1, value=scenario['id'])
                ws_scenarios.cell(row=row_idx, column=2, value=scenario['title'])
                ws_scenarios.cell(row=row_idx, column=3, value=scenario['priority'])
                ws_scenarios.cell(row=row_idx, column=4, value=scenario['test_type'])
                ws_scenarios.cell(row=row_idx, column=5, value=scenario['given'])
                ws_scenarios.cell(row=row_idx, column=6, value=scenario['when'])
                ws_scenarios.cell(row=row_idx, column=7, value=scenario['then'])
            
            # Adjust column widths
            ws_scenarios.column_dimensions['A'].width = 12
            ws_scenarios.column_dimensions['B'].width = 35
            ws_scenarios.column_dimensions['C'].width = 10
            ws_scenarios.column_dimensions['D'].width = 10
            ws_scenarios.column_dimensions['E'].width = 30
            ws_scenarios.column_dimensions['F'].width = 30
            ws_scenarios.column_dimensions['G'].width = 30
            
            # Save workbook
            wb.save(str(filepath))
            logger.info(f"Excel test plan generated successfully: {filepath}")

            return str(filepath)

        except Exception as e:
            logger.error(f"Failed to generate Excel test plan: {str(e)}")
            raise

    def generate_both(self, test_plan: TestPlan, jira_ticket_key: str) -> Dict[str, str]:
        """
        Generate both PDF and Excel documents

        Args:
            test_plan: TestPlan object
            jira_ticket_key: Jira ticket key

        Returns:
            Dictionary with paths to both files
        """
        logger.info(f"Generating both PDF and Excel test plans for {jira_ticket_key}")
        
        pdf_path = self.generate_pdf(test_plan, jira_ticket_key)
        excel_path = self.generate_excel(test_plan, jira_ticket_key)

        return {
            'pdf': pdf_path,
            'excel': excel_path
        }