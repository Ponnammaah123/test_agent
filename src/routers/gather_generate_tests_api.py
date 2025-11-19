import re
import io
import requests
import openpyxl
from fastapi import APIRouter, HTTPException, Body, status
from pydantic import BaseModel, Field
from typing import Dict, Any, List, Optional, Tuple

# Import core clients and config
from src.config.settings import Config
from src.clients.jira_client import JiraClient
from src.clients.github_client import GitHubClient
from src.utils.logger import get_logger
from src.utils.exceptions import JiraClientException, GitHubClientException

# Import the agent and models needed for generation
from src.agents.test_generation_agent import TestGenerationAgent
from src.models.test_plan_models import TestPlan
from src.models.github_models import CodebaseAnalysis

# Import TestRepoClient for the push/PR operation
from src.clients.test_repo_client import TestRepoClient

logger = get_logger(__name__)
router = APIRouter()

# --- Pydantic Models ---

class GatherGenerateTestsRequest(BaseModel):
    """Input model for the new orchestration endpoint"""
    jira_ticket_key: str = Field(..., description="Jira ticket key (e.g., QEA-19)")
    scope_analysis: Dict[str, Any] = Field(..., description="Scope analysis provided by the user")


# --- Helper Functions ---

def _parse_comment_data(comments: List[Dict[str, Any]]) -> Tuple[str, str]:
    """
    Parses comments to find the latest branch and test repo URL.
    """
    logger.info(f"Parsing {len(comments)} comments for branch and repo URL...")
    
    # Regex patterns to find data in comments
    repo_pattern = re.compile(r"\*Test Repository:\*.*?\[(https://.*?)\|")
    branch_pattern = re.compile(r"\*Branch:\*.*?{monospace}(.*?){monospace}")
    
    # Iterate in reverse (newest comments first)
    for comment in reversed(comments):
        body = comment.get('body', '')
        if "E2E Tests Generated Successfully" in body:
            logger.info("Found 'E2E Tests Generated' comment.")
            try:
                repo_url_match = repo_pattern.search(body)
                branch_match = branch_pattern.search(body)
                
                if repo_url_match and branch_match:
                    repo_url = repo_url_match.group(1)
                    branch = branch_match.group(1)
                    logger.info(f"Successfully parsed: Branch='{branch}', Repo='{repo_url}'")
                    return branch, repo_url
            except Exception as e:
                logger.warning(f"Failed to parse comment body: {e}")
                continue
                
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail="Couldadaşs not find a comment with 'E2E Tests Generated Successfully' containing branch and repo URL."
    )

def _parse_excel_test_plan(attachments: List[Dict[str, Any]], config: Config) -> Dict[str, Any]:
    """
    Downloads and parses the .xlsx test plan attachment.
    
    Args:
        attachments: List of attachment metadata from Jira.
        config: The application configuration object (for auth).
    """
    logger.info(f"Searching for .xlsx test plan in {len(attachments)} attachments...")
    
    excel_url = None
    for attachment in attachments:
        filename = attachment.get('filename', '')
        if "TestPlan" in filename and filename.endswith(".xlsx"):
            excel_url = attachment.get('content')
            logger.info(f"Found Excel test plan: {filename}")
            break
            
    if not excel_url:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Could not find a '.xlsx' test plan attachment."
        )
        
    try:
        # Download the Excel file
        logger.info(f"Downloading Excel file from: {excel_url}")
        
        # Jira attachments require the same auth as the API.
        jira_auth = (config.jira.email, config.jira.api_token)
        response = requests.get(excel_url, auth=jira_auth)
        
        response.raise_for_status()
        
        # Load into memory
        file_bytes = io.BytesIO(response.content)
        wb = openpyxl.load_workbook(file_bytes, data_only=True)
        
        # --- Parse "Overview" Sheet ---
        overview_data = {}
        ws_overview = wb.get_sheet_by_name("Overview")
        for row in ws_overview.iter_rows(min_row=3, max_col=2):
            key = row[0].value
            value = row[1].value
            if key:
                overview_data[key.lower().replace(' ', '_')] = value
        
        # --- Parse "Test Scenarios" Sheet ---
        scenarios = []
        ws_scenarios = wb.get_sheet_by_name("Test Scenarios")
        
        # Get headers
        headers = [cell.value.lower() for cell in ws_scenarios[1]]
        
        for row in ws_scenarios.iter_rows(min_row=2):
            scenario_data = {headers[i]: cell.value for i, cell in enumerate(row)}
            if scenario_data.get('id'): # Only add if row has data
                scenarios.append({
                    "id": scenario_data.get('id'),
                    "title": scenario_data.get('title'),
                    "priority": scenario_data.get('priority'),
                    "test_type": scenario_data.get('type'), # Map 'type' from Excel to 'test_type'
                    "given": scenario_data.get('given'),
                    "when": scenario_data.get('when'),
                    "then": scenario_data.get('then'),
                })
        
        logger.info(f"Parsed {len(scenarios)} test scenarios from Excel.")
        
        # Assemble the test_plan object
        test_plan = {
            "jira_ticket": overview_data.get('jira_ticket', ''),
            "strategy": overview_data.get('strategy', ''),
            "test_approach": overview_data.get('test_approach', ''),
            "confidence_score": overview_data.get('confidence_score', 0.0),
            "test_scenarios": scenarios,
        }
        
        return test_plan

    except requests.RequestException as e:
        logger.error(f"Failed to download Excel file: {e}")
        error_detail = f"Failed to download attachment: {e}"
        if e.response is not None:
             error_detail = f"Failed to download attachment: {e.response.status_code} {e.response.reason} for url: {excel_url}"
        raise HTTPException(status_code=502, detail=error_detail)
    except Exception as e:
        logger.error(f"Failed to parse Excel file: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to parse Excel file: {e}")

# --- API Endpoint ---

@router.post(
    "/gather_response_for_generate/tests",
    response_model=Dict[str, Any],
    summary="Orchestrate and Generate Tests",
    description="Gathers all data AND generates the test code in a single call."
)
async def gather_response_for_generate_tests(
    request: GatherGenerateTestsRequest = Body(...)
) -> Dict[str, Any]:
    """
    This endpoint automates the entire data gathering and generation service.
    """
    
    try:
        config = Config()
        config.validate()
        
        jira_ticket_key = request.jira_ticket_key
        logger.info(f"[{jira_ticket_key}] Starting full test generation orchestration...")
        
        # --- Step 1: Fetch Jira Data ---
        jira_client = JiraClient(config)
        
        logger.info(f"[{jira_ticket_key}] (1/6) Fetching comments and attachments...")
        comments = jira_client.get_comments(jira_ticket_key, top_n=10)
        attachments = jira_client.get_attachments(jira_ticket_key)
        
        if not comments or not attachments:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No comments or attachments found for this ticket. Agent may not have run."
            )

        # --- Step 2: Parse Comments for Test Repo URL ---
        logger.info(f"[{jira_ticket_key}] (2/6) Parsing comment data...")
        test_branch_from_jira, test_repo_url = _parse_comment_data(comments)

        # --- Step 3: Parse Test Plan from Excel Attachment ---
        logger.info(f"[{jira_ticket_key}] (3/6) Parsing test plan from Excel attachment...")
        test_plan = _parse_excel_test_plan(attachments, config)
        
        # --- Step 4: Get Codebase Analysis (App Repo Scan) ---
        target_app_branch = "main" 
        
        logger.info(f"[{jira_ticket_key}] (4/6) Fetching ALL files from APP repo branch: {target_app_branch}...")
        github_client = GitHubClient(config)
        
        all_files_content = github_client.get_all_files_in_branch(target_app_branch)
        
        files_list = []
        for path, content in all_files_content.items():
            files_list.append({
                "path": path,
                "status": "existing",
                "additions": 0,
                "deletions": 0,
                "content": content 
            })
            
            # CRITICAL: Pre-populate the cache for the agent to find the content
            if github_client.cache:
                from src.clients.github_client_cache import CachedFile, CachedAnalysis
                import datetime
                
                cached_file = CachedFile(
                    path=path,
                    status="existing",
                    content=content,
                    file_size_bytes=len(content) if content else 0
                )
                
                cache_key = f"{github_client.repo}:{target_app_branch}"
                if cache_key not in github_client.cache.cache:
                     analysis = CachedAnalysis(
                         repository=github_client.repo,
                         branch=target_app_branch,
                         commit_id="latest",
                         files={}
                     )
                     github_client.cache.cache[cache_key] = analysis
                     github_client.cache.access_times[cache_key] = datetime.datetime.now()
                
                github_client.cache.cache[cache_key].files[path] = cached_file

        logger.info(f"[{jira_ticket_key}] Successfully fetched and cached {len(files_list)} files.")

        codebase_analysis_obj = CodebaseAnalysis(
            repository=config.github.repo,
            branch=target_app_branch,
            test_coverage=0.0,
            files_changed=files_list, 
            components_identified=[],
            commit_count=1
        )

        # --- Step 5: Generate Test Code ---
        logger.info(f"[{jira_ticket_key}] (5/6) Running Test Generation Agent...")
        
        agent = TestGenerationAgent(config)

        try:
            test_plan_obj = TestPlan(**test_plan)
        except Exception as e:
            logger.error(f"Failed to parse models for agent: {e}")
            raise HTTPException(status_code=400, detail=f"Invalid data for agent models: {e}")

        test_generation_result = await agent.generate_tests(
            jira_ticket_key=jira_ticket_key,
            test_plan=test_plan_obj,
            codebase_analysis=codebase_analysis_obj, 
            test_repo_url=test_repo_url,
            scope_analysis=request.scope_analysis
        )
        
        generated_files = test_generation_result.get('test_files', {})
        if not generated_files:
            return {
                "status": "success",
                "message": "No new tests were generated (possibly due to duplicates or empty scenarios). Skipping PR.",
                "data": test_generation_result
            }

        # --- Step 6: Commit and Raise PR ---
        logger.info(f"[{jira_ticket_key}] (6/6) Committing {len(generated_files)} files and raising PR...")
        
        test_repo_client = TestRepoClient(config, test_repo_url)
        
        # Push to the test repo's main branch, creating a new feature branch for the PR.
        pr_result = test_repo_client.push_tests_and_create_pr(
            files_to_commit=generated_files, 
            jira_ticket_key=jira_ticket_key,
            base_branch="main" # PR target
        )

        final_response = {
            "status": "success",
            "message": f"Generated {len(generated_files)} test files and created PR.",
            "pr_url": pr_result.get('pr_url'),
            "pr_branch": pr_result.get('new_branch_name'),
            "data": test_generation_result
        }

        logger.info(f"[{jira_ticket_key}] Orchestration complete. PR URL: {pr_result.get('pr_url')}")
        return final_response
    
    except HTTPException:
        raise 
    except (JiraClientException, GitHubClientException) as e:
        logger.error(f"[{request.jira_ticket_key}] Client Error: {e}")
        raise HTTPException(status_code=502, detail=f"Client Error: {e}")
    except Exception as e:
        logger.error(f"[{request.jira_ticket_key}] Unexpected Error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Unexpected Internal Error: {e}")